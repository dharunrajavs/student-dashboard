from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from app import db
from app.models.user import User
from app.models.student import Student
from app.models.performance import Performance, Alert
from sqlalchemy import func
from app.services.risk_scorer import risk_scoring_service
from app.services.scholarship_predictor import scholarship_predictor_service
from datetime import datetime, timedelta
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

bp = Blueprint('admin', __name__)

@bp.route('/dashboard', methods=['GET'])
@jwt_required()
def get_admin_dashboard():
    """
    Get admin dashboard analytics
    ---
    tags:
      - Admin
    security:
      - Bearer: []
    responses:
      200:
        description: Dashboard analytics
      403:
        description: Unauthorized
    """
    try:
        # Verify admin role
        user_id = int(get_jwt_identity())
        user = User.query.get(user_id)
        
        if user.role not in ['admin', 'faculty']:
            return jsonify({'error': 'Unauthorized'}), 403
        
        # Total students
        total_students = Student.query.count()
        
        # Department-wise distribution
        dept_distribution = db.session.query(
            Student.department,
            func.count(Student.id).label('count')
        ).group_by(Student.department).all()
        
        # At-risk students (low GPA)
        at_risk_students = Student.query.filter(Student.current_cgpa < 6.0).count()
        
        # Excellent performers
        excellent_students = Student.query.filter(Student.current_cgpa >= 8.5).count()
        
        # Average CGPA
        avg_cgpa = db.session.query(func.avg(Student.current_cgpa)).scalar() or 0
        
        # Recent alerts
        recent_alerts = Alert.query.filter_by(
            is_read=False
        ).order_by(Alert.created_at.desc()).limit(10).all()
        
        # Performance trends (last 30 days)
        thirty_days_ago = datetime.utcnow() - timedelta(days=30)
        recent_performances = Performance.query.filter(
            Performance.created_at >= thirty_days_ago
        ).all()
        
        avg_recent_marks = sum(p.total_marks for p in recent_performances) / len(recent_performances) if recent_performances else 0
        avg_recent_attendance = sum(p.attendance_percentage for p in recent_performances) / len(recent_performances) if recent_performances else 0
        
        return jsonify({
            'total_students': total_students,
            'department_distribution': [
                {'department': dept, 'count': count}
                for dept, count in dept_distribution
            ],
            'at_risk_students': at_risk_students,
            'excellent_students': excellent_students,
            'average_cgpa': round(avg_cgpa, 2),
            'recent_alerts': [a.to_dict() for a in recent_alerts],
            'performance_overview': {
                'average_marks': round(avg_recent_marks, 2),
                'average_attendance': round(avg_recent_attendance, 2),
                'total_records': len(recent_performances)
            }
        }), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/students/at-risk', methods=['GET'])
@jwt_required()
def get_at_risk_students():
    """
    Get list of at-risk students
    ---
    tags:
      - Admin
    security:
      - Bearer: []
    parameters:
      - in: query
        name: threshold
        type: integer
        default: 70
    responses:
      200:
        description: At-risk students list
    """
    try:
        user_id = int(get_jwt_identity())
        user = User.query.get(user_id)
        
        if user.role not in ['admin', 'faculty']:
            return jsonify({'error': 'Unauthorized'}), 403
        
        risk_threshold = request.args.get('threshold', 70, type=int)
        
        students = Student.query.all()
        at_risk_students = []
        
        for student in students:
            risk_data = risk_scoring_service.calculate_risk_score(student.id)
            
            if risk_data and risk_data.get('risk_score', 0) >= risk_threshold:
                at_risk_students.append({
                    'student': student.to_dict(),
                    'risk_data': risk_data
                })
        
        # Sort by risk score (highest first)
        at_risk_students.sort(key=lambda x: x['risk_data']['risk_score'], reverse=True)
        
        return jsonify({
            'at_risk_students': at_risk_students,
            'total': len(at_risk_students),
            'threshold': risk_threshold
        }), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/scholarships/eligible', methods=['GET'])
@jwt_required()
def get_eligible_students():
    """
    Get scholarship eligible students
    ---
    tags:
      - Admin
    security:
      - Bearer: []
    responses:
      200:
        description: Eligible students list
    """
    try:
        user_id = int(get_jwt_identity())
        user = User.query.get(user_id)
        
        if user.role not in ['admin']:
            return jsonify({'error': 'Unauthorized'}), 403
        
        students = Student.query.all()
        eligible_students = []
        
        for student in students:
            # Get latest performance
            latest_perf = Performance.query.filter_by(
                student_id=student.id
            ).order_by(Performance.created_at.desc()).first()
            
            if latest_perf:
                data = {
                    'gpa': student.current_cgpa,
                    'attendance_percentage': latest_perf.attendance_percentage,
                    'family_income': student.family_income or 0,
                    'extracurricular_score': latest_perf.participation_score or 0,
                    'discipline_score': latest_perf.discipline_score or 100,
                    'research_publications': 0
                }
                
                result = scholarship_predictor_service.predict(data)
                
                if result.get('eligible'):
                    eligible_students.append({
                        'student': student.to_dict(),
                        'eligibility': result
                    })
        
        # Sort by probability
        eligible_students.sort(key=lambda x: x['eligibility']['probability'], reverse=True)
        
        return jsonify({
            'eligible_students': eligible_students,
            'total': len(eligible_students)
        }), 200
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/scholarships/export', methods=['GET'])
@jwt_required()
def export_scholarship_eligible():
    """
    Export eligible students to Excel
    ---
    tags:
      - Admin
    security:
      - Bearer: []
    responses:
      200:
        description: Excel file
    """
    try:
        user_id = int(get_jwt_identity())
        user = User.query.get(user_id)
        
        if user.role not in ['admin']:
            return jsonify({'error': 'Unauthorized'}), 403
        
        students = Student.query.all()
        eligible_students = []
        
        for student in students:
            latest_perf = Performance.query.filter_by(
                student_id=student.id
            ).order_by(Performance.created_at.desc()).first()
            
            if latest_perf:
                data = {
                    'gpa': student.current_cgpa,
                    'attendance_percentage': latest_perf.attendance_percentage,
                    'family_income': student.family_income or 0,
                    'extracurricular_score': latest_perf.participation_score or 0,
                    'discipline_score': latest_perf.discipline_score or 100,
                    'research_publications': 0
                }
                
                result = scholarship_predictor_service.predict(data)
                
                if result.get('eligible'):
                    eligible_students.append({
                        'roll_number': student.roll_number,
                        'name': student.full_name,
                        'department': student.department,
                        'cgpa': student.current_cgpa,
                        'probability': result['probability']
                    })
        
        # Create Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "Scholarship Eligible Students"
        
        # Headers
        headers = ['Roll Number', 'Name', 'Department', 'CGPA', 'Eligibility Probability (%)']
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for row, student in enumerate(eligible_students, 2):
            ws.cell(row=row, column=1, value=student['roll_number'])
            ws.cell(row=row, column=2, value=student['name'])
            ws.cell(row=row, column=3, value=student['department'])
            ws.cell(row=row, column=4, value=student['cgpa'])
            ws.cell(row=row, column=5, value=student['probability'])
        
        # Auto-adjust column widths
        for col in range(1, 6):
            ws.column_dimensions[chr(64 + col)].width = 20
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'scholarship_eligible_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@bp.route('/alerts/create', methods=['POST'])
@jwt_required()
def create_alert():
    """
    Create an alert for a student
    ---
    tags:
      - Admin
    security:
      - Bearer: []
    parameters:
      - in: body
        name: body
        required: true
        schema:
          type: object
    responses:
      201:
        description: Alert created
    """
    try:
        user_id = int(get_jwt_identity())
        user = User.query.get(user_id)
        
        if user.role not in ['admin', 'faculty']:
            return jsonify({'error': 'Unauthorized'}), 403
        
        data = request.get_json()
        
        alert = Alert(
            student_id=data['student_id'],
            alert_type=data.get('alert_type', 'general'),
            severity=data.get('severity', 'medium'),
            title=data['title'],
            message=data['message']
        )
        
        db.session.add(alert)
        db.session.commit()
        
        return jsonify({
            'message': 'Alert created successfully',
            'alert': alert.to_dict()
        }), 201
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@bp.route('/students/<int:student_id>/deactivate', methods=['POST'])
@jwt_required()
def deactivate_student(student_id):
    """
    Deactivate a student account
    ---
    tags:
      - Admin
    security:
      - Bearer: []
    parameters:
      - in: path
        name: student_id
        type: integer
    responses:
      200:
        description: Student deactivated
    """
    try:
        user_id = int(get_jwt_identity())
        user = User.query.get(user_id)
        
        if user.role not in ['admin']:
            return jsonify({'error': 'Unauthorized'}), 403
        
        student = Student.query.get(student_id)
        if not student:
            return jsonify({'error': 'Student not found'}), 404
        
        student.user.is_active = False
        db.session.commit()
        
        return jsonify({'message': 'Student account deactivated'}), 200
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500
